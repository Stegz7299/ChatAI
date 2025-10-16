from PyPDF2 import PdfReader
import docx
from openpyxl import load_workbook

def read_file(path: str) -> str:

    text = ""

    try:
        if path.endswith(".pdf"):
            reader = PdfReader(path)
            text = "\n".join([p.extract_text() or "" for p in reader.pages])

        elif path.endswith(".docx"):
            doc = docx.Document(path)
            text = "\n".join([p.text for p in doc.paragraphs])

        elif path.endswith(".txt"):
            with open(path, "r", encoding="utf-8") as f:
                text = f.read()

        elif path.endswith(".xlsx"):
            wb = load_workbook(path, data_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        rows.append(row_text)
            text = "\n".join(rows)

        else:
            text = f"[Unsupported file type: {path}]"

    except Exception as e:
        text = f"[Error reading file {path}: {str(e)}]"

    return text